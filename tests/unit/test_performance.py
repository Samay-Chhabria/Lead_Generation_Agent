"""Performance and stability tests.

Verifies the pipeline stays fast and leak-free at realistic and larger lead
volumes. Time bounds are deliberately generous so slow machines do not flake;
their purpose is to catch order-of-magnitude regressions, not to micro-benchmark.
"""

import time

from openpyxl import load_workbook

from app.models.lead import Lead
from app.pipeline.application_pipeline import ApplicationPipeline
from app.processing.processing_pipeline import ProcessingPipeline

PROMPT = "software companies in Karachi"
_PROCESSING_BUDGET_SECONDS = 3.0
_FULL_RUN_BUDGET_SECONDS = 8.0


def _lead(index: int) -> Lead:
    return Lead(
        business_name=f"Business {index}",
        phone_number="+10000000000",
        email="",
        website=f"https://site{index}.example",
        location="Karachi",
        provider="fixed",
        search_query=PROMPT,
        source_url=f"https://maps.example/place/{index}",
    )


def test_processing_fifty_leads_within_budget() -> None:
    leads = [_lead(index) for index in range(50)]

    started = time.perf_counter()
    result = ProcessingPipeline().process(leads)
    elapsed = time.perf_counter() - started

    assert result.final_count == 50
    assert result.duplicates_removed == 0
    assert elapsed < _PROCESSING_BUDGET_SECONDS


def test_processing_one_hundred_leads_within_budget() -> None:
    leads = [_lead(index) for index in range(100)]

    started = time.perf_counter()
    result = ProcessingPipeline().process(leads)
    elapsed = time.perf_counter() - started

    assert result.final_count == 100
    assert elapsed < _PROCESSING_BUDGET_SECONDS


def test_full_run_with_fifty_leads_within_budget(fixed_settings, fixed_factory) -> None:
    leads = [_lead(index) for index in range(50)]

    result = ApplicationPipeline(settings=fixed_settings, factory=fixed_factory(leads)).execute(
        PROMPT
    )

    assert result.success is True
    assert result.collected_leads == 50
    assert result.processed_leads == 50
    assert result.execution_time < _FULL_RUN_BUDGET_SECONDS


def test_full_run_with_one_hundred_leads_exports_all_rows(fixed_settings, fixed_factory) -> None:
    leads = [_lead(index) for index in range(100)]

    result = ApplicationPipeline(settings=fixed_settings, factory=fixed_factory(leads)).execute(
        PROMPT
    )

    assert result.success is True
    assert result.collected_leads == 100
    assert result.excel_output_path is not None
    workbook = load_workbook(result.excel_output_path)
    assert workbook["Leads"].max_row == 101


def test_repeated_runs_close_browser_each_time(fixed_settings, fixed_factory, browser) -> None:
    for _ in range(3):
        result = ApplicationPipeline(
            settings=fixed_settings, factory=fixed_factory([_lead(0)])
        ).execute(PROMPT)
        assert result.success is True

    assert browser.close_count == 3


def test_repeated_runs_do_not_bleed_results_between_runs(fixed_settings, fixed_factory) -> None:
    first = ApplicationPipeline(
        settings=fixed_settings, factory=fixed_factory([_lead(0), _lead(1)])
    ).execute(PROMPT)
    second = ApplicationPipeline(settings=fixed_settings, factory=fixed_factory([])).execute(PROMPT)

    assert first.collected_leads == 2
    assert second.collected_leads == 0
