"""Tests for the Excel export pipeline (Milestone 7)."""

import logging
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.exceptions.export_exception import ExportException
from app.exporter import ExcelExporter, FileManager, WorkbookBuilder
from app.exporter.workbook_builder import COLUMN_HEADERS
from app.models.lead import Lead

FIXED_TIME = datetime(2026, 8, 4, 12, 30, 45)


def _lead(**overrides) -> Lead:
    fields = {"business_name": "Acme Corp"}
    fields.update(overrides)
    return Lead(**fields)


def _make_lead() -> Lead:
    return _lead(
        business_name="Acme Corp",
        email="info@acme.example",
        phone_number="+12125551234",
        website="https://acme.example",
        location="New York",
        provider="google",
        search_query="coffee shops in America",
        source_url="https://maps.example/place/acme",
        collected_at=FIXED_TIME,
    )


class TestFileManager:
    def test_generate_filename_builds_meaningful_name(self, tmp_path: Path) -> None:
        manager = FileManager(tmp_path)

        assert (
            manager.generate_filename("coffee shops", "america")
            == "leads_coffee_shops_america.xlsx"
        )

    def test_generate_filename_omits_empty_location(self, tmp_path: Path) -> None:
        manager = FileManager(tmp_path)

        assert manager.generate_filename("dentists") == "leads_dentists.xlsx"

    def test_generate_filename_sanitizes_components(self, tmp_path: Path) -> None:
        manager = FileManager(tmp_path)

        name = manager.generate_filename("  Software   Companies ", "KARACHI/")

        assert name == "leads_Software_Companies_KARACHI.xlsx"

    def test_generate_filename_rejects_empty_business_type(self, tmp_path: Path) -> None:
        manager = FileManager(tmp_path)

        with pytest.raises(ExportException, match="Business type"):
            manager.generate_filename("///")

    def test_resolve_path_uses_base_filename_when_free(self, tmp_path: Path) -> None:
        manager = FileManager(tmp_path)

        path = manager.resolve_path("coffee shops", "america")

        assert path == tmp_path / "leads_coffee_shops_america.xlsx"

    def test_resolve_path_appends_timestamp_on_collision(self, tmp_path: Path) -> None:
        manager = FileManager(tmp_path, clock=lambda: FIXED_TIME)
        (tmp_path / "leads_coffee_shops_america.xlsx").write_bytes(b"existing")

        path = manager.resolve_path("coffee shops", "america")

        assert path == tmp_path / "leads_coffee_shops_america_20260804_123045.xlsx"

    def test_save_path_creates_output_directory(self, tmp_path: Path) -> None:
        output_dir = tmp_path / "outputs" / "nested"
        manager = FileManager(output_dir)

        path = manager.save_path("coffee shops", "america")

        assert path.parent == output_dir
        assert output_dir.is_dir()

    def test_save_path_raises_when_directory_cannot_be_created(self, tmp_path: Path) -> None:
        blocker = tmp_path / "not_a_directory"
        blocker.write_bytes(b"blocking file")
        manager = FileManager(blocker)

        with pytest.raises(ExportException, match="output directory"):
            manager.save_path("coffee shops")


class TestWorkbookBuilder:
    def test_build_creates_leads_sheet_with_headers(self) -> None:
        workbook = WorkbookBuilder().build([_make_lead()])

        sheet = workbook["Leads"]
        assert [cell.value for cell in sheet[1]] == COLUMN_HEADERS

    def test_build_writes_one_row_per_lead(self) -> None:
        workbook = WorkbookBuilder().build([_make_lead(), _make_lead()])

        sheet = workbook["Leads"]
        assert sheet.max_row == 3
        assert sheet["A2"].value == "Acme Corp"
        assert sheet["B2"].value == "info@acme.example"
        assert sheet["C2"].value == "+12125551234"
        assert sheet["D2"].value == "https://acme.example"
        assert sheet["E2"].value == "New York"
        assert sheet["F2"].value == "google"
        assert sheet["G2"].value == "coffee shops in America"
        assert sheet["H2"].value == "2026-08-04 12:30:45"
        assert sheet["I2"].value == "https://maps.example/place/acme"

    def test_build_formats_headers_and_freezes_pane(self) -> None:
        workbook = WorkbookBuilder().build([_make_lead()])

        sheet = workbook["Leads"]
        assert sheet.freeze_panes == "A2"
        assert all(cell.font.bold for cell in sheet[1])

    def test_build_writes_empty_fields_as_blank_cells(self) -> None:
        workbook = WorkbookBuilder().build([_lead(business_name="Minimal Co")])

        sheet = workbook["Leads"]
        assert sheet["B2"].value == ""
        assert sheet["C2"].value == ""
        assert sheet["D2"].value == ""
        assert sheet["E2"].value == ""
        assert sheet["F2"].value == ""
        assert sheet["G2"].value == ""
        assert sheet["I2"].value == ""

    def test_build_preserves_unicode(self) -> None:
        workbook = WorkbookBuilder().build(
            [_lead(business_name="Café Zürich", email="über@cafe.example")]
        )

        sheet = workbook["Leads"]
        assert sheet["A2"].value == "Café Zürich"
        assert sheet["B2"].value == "über@cafe.example"

    def test_build_autosizes_columns(self) -> None:
        workbook = WorkbookBuilder().build([_make_lead()])

        sheet = workbook["Leads"]
        assert sheet.column_dimensions["A"].width >= len("Business Name")
        assert sheet.column_dimensions["D"].width >= len("https://acme.example")
        assert sheet.column_dimensions["I"].width >= len("https://maps.example/place/acme")

    def test_build_empty_leads_produces_headers_only(self) -> None:
        workbook = WorkbookBuilder().build([])

        sheet = workbook["Leads"]
        assert sheet.max_row == 1


class TestExcelExporter:
    def test_export_saves_a_readable_workbook(self, tmp_path: Path) -> None:
        exporter = ExcelExporter(file_manager=FileManager(tmp_path))
        leads = [_make_lead()]

        path = exporter.export(leads, "coffee shops", "america")

        assert path == tmp_path / "leads_coffee_shops_america.xlsx"
        assert path.exists()
        reloaded = load_workbook(path)
        sheet = reloaded["Leads"]
        assert sheet["A1"].value == "Business Name"
        assert sheet["A2"].value == "Acme Corp"
        assert reloaded.sheetnames == ["Leads"]

    def test_export_empty_leads_writes_headers_only(self, tmp_path: Path) -> None:
        exporter = ExcelExporter(file_manager=FileManager(tmp_path))

        path = exporter.export([], "coffee shops", "america")

        reloaded = load_workbook(path)
        assert reloaded["Leads"].max_row == 1

    def test_export_does_not_overwrite_existing_workbook(self, tmp_path: Path) -> None:
        manager = FileManager(tmp_path, clock=lambda: FIXED_TIME)
        exporter = ExcelExporter(file_manager=manager)
        (tmp_path / "leads_coffee_shops_america.xlsx").write_bytes(b"keep me")
        exporter.export([_make_lead()], "coffee shops", "america")

        first = tmp_path / "leads_coffee_shops_america.xlsx"
        second = tmp_path / "leads_coffee_shops_america_20260804_123045.xlsx"
        assert first.read_bytes() == b"keep me"
        assert second.exists()

    def test_export_uses_injected_workbook_builder(self, tmp_path: Path) -> None:
        class TrackingBuilder(WorkbookBuilder):
            def __init__(self) -> None:
                super().__init__()
                self.built_with: list[Lead] | None = None

            def build(self, leads: list[Lead]):
                self.built_with = leads
                return super().build(leads)

        builder = TrackingBuilder()
        leads = [_make_lead()]
        exporter = ExcelExporter(workbook_builder=builder, file_manager=FileManager(tmp_path))

        exporter.export(leads, "coffee shops", "america")

        assert builder.built_with == leads

    def test_export_wraps_save_failures(self, tmp_path: Path) -> None:
        class FailingBuilder(WorkbookBuilder):
            def build(self, leads: list[Lead]):
                workbook = super().build(leads)

                def fail(path: Path) -> None:
                    raise PermissionError("access denied")

                workbook.save = fail  # type: ignore[method-assign]
                return workbook

        exporter = ExcelExporter(
            workbook_builder=FailingBuilder(), file_manager=FileManager(tmp_path)
        )

        with pytest.raises(ExportException, match="could not be saved"):
            exporter.export([_make_lead()], "coffee shops", "america")

    def test_export_logs_lifecycle(self, tmp_path: Path, caplog: pytest.LogCaptureFixture) -> None:
        exporter = ExcelExporter(file_manager=FileManager(tmp_path))

        with caplog.at_level(logging.INFO):
            exporter.export([_make_lead(), _make_lead()], "coffee shops", "america")

        messages = [record.message for record in caplog.records]
        assert any("Exporting leads" in message for message in messages)
        assert any("Workbook created" in message for message in messages)
        assert any("Rows written: 2" in message for message in messages)
        assert any("Workbook saved" in message for message in messages)
        assert any("Export completed" in message for message in messages)
        assert any("Output:" in message for message in messages)
