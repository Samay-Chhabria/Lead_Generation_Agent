"""End-to-end tests running the complete user-facing workflow.

The full workflow — prompt parsing, provider search, lead collection,
processing, Excel export, and execution summary — runs through the real
ApplicationPipeline with the fake 'fixed' provider, so every requirement is
verified without a browser or network (Requirements 1, 11, 12).
"""

import logging

import pytest
from openpyxl import load_workbook

from app.exporter.workbook_builder import COLUMN_HEADERS
from app.models.lead import Lead
from app.pipeline.application_pipeline import ApplicationPipeline
from tests.fakes import FakeBrowser, FakeElement, FakePage, build_fixed_factory

PROMPT = "software companies in Karachi"
MAILTO_SELECTOR = 'a[href^="mailto:"]'


def _lead(
    name: str = "Acme Corp",
    *,
    phone: str = "",
    email: str = "",
    website: str = "",
    location: str = "",
) -> Lead:
    return Lead(
        business_name=name,
        phone_number=phone,
        email=email,
        website=website,
        location=location,
        provider="fixed",
        search_query=PROMPT,
        source_url=f"https://maps.example/place/{name}",
    )


def _run(fixed_settings, fixed_factory, leads: list[Lead]):
    return ApplicationPipeline(settings=fixed_settings, factory=fixed_factory(leads)).execute(
        PROMPT
    )


# --- Successful runs ---------------------------------------------------------


def test_full_run_exports_workbook_with_required_columns_and_leads(
    fixed_settings, fixed_factory
) -> None:
    leads = [
        _lead(
            "Alpha Corp",
            phone="+12125550001",
            email="a@alpha.example",
            website="https://alpha.example",
        ),
        _lead("Beta Ltd", location="Karachi"),
    ]

    result = _run(fixed_settings, fixed_factory, leads)

    assert result.success is True
    assert result.business_type == "software companies"
    assert result.location == "Karachi"
    assert result.collected_leads == 2
    assert result.excel_output_path is not None
    assert result.excel_output_path.exists()

    sheet = load_workbook(result.excel_output_path)["Leads"]
    header_row = [cell.value for cell in sheet[1]]
    assert header_row == COLUMN_HEADERS
    assert sheet.max_row == 3
    assert sheet["A2"].value == "Alpha Corp"
    assert sheet["B2"].value == "a@alpha.example"
    assert sheet["D2"].value == "https://alpha.example"
    assert sheet["A3"].value == "Beta Ltd"


def test_full_run_prints_success_summary(
    fixed_settings, fixed_factory, capsys: pytest.CaptureFixture
) -> None:
    _run(fixed_settings, fixed_factory, [_lead("Alpha Corp")])

    output = capsys.readouterr().out
    assert "Search Plan" in output
    assert "Original Prompt: software companies in Karachi" in output
    assert "Lead Generation Completed Successfully" in output
    assert "Businesses Found: 1" in output
    assert "Leads Exported: 1" in output
    assert "Execution Time:" in output


def test_full_run_closes_browser(fixed_settings, fixed_factory, browser) -> None:
    _run(fixed_settings, fixed_factory, [_lead("Alpha Corp")])

    assert browser.close_count == 1
    assert browser.is_running() is False


def test_full_run_logs_every_stage(
    fixed_settings, fixed_factory, caplog: pytest.LogCaptureFixture
) -> None:
    with caplog.at_level(logging.INFO):
        _run(fixed_settings, fixed_factory, [_lead("Alpha Corp")])

    messages = [record.message for record in caplog.records]
    for stage in (
        "Pipeline started.",
        "Parsed search plan:",
        "Processing completed.",
        "Excel exported.",
        "Summary generated.",
        "Pipeline finished.",
    ):
        assert any(stage in message for message in messages)


def test_full_run_without_results_still_exports_empty_workbook(
    fixed_settings, fixed_factory
) -> None:
    result = _run(fixed_settings, fixed_factory, [])

    assert result.success is True
    assert result.collected_leads == 0
    assert result.excel_output_path is not None
    sheet = load_workbook(result.excel_output_path)["Leads"]
    assert [cell.value for cell in sheet[1]] == COLUMN_HEADERS
    assert sheet.max_row == 1


def test_full_run_deduplicates_and_drops_invalid_leads(fixed_settings, fixed_factory) -> None:
    leads = [
        _lead("Acme Corp", website="https://acme.example"),
        _lead("Acme Corp", website="https://acme.example"),
        _lead("   "),
        _lead("Beta Ltd"),
    ]

    result = _run(fixed_settings, fixed_factory, leads)

    assert result.success is True
    assert result.collected_leads == 4
    assert result.duplicates_removed == 1
    assert result.processed_leads == 2
    sheet = load_workbook(result.excel_output_path)["Leads"]
    assert sheet.max_row == 3
    assert sheet["A2"].value == "Acme Corp"
    assert sheet["A3"].value == "Beta Ltd"


def test_full_run_survives_website_failures(fixed_settings) -> None:
    page = FakePage(
        elements_by_url={
            "https://good.example": {
                MAILTO_SELECTOR: [FakeElement(attributes={"href": "mailto:hello@good.example"})]
            }
        },
        goto_errors={"https://broken.example": RuntimeError("connection refused")},
    )
    browser = FakeBrowser(page)
    factory = build_fixed_factory(
        fixed_settings,
        browser,
        [
            _lead("Good", website="https://good.example"),
            _lead("Broken", website="https://broken.example"),
        ],
    )

    result = ApplicationPipeline(settings=fixed_settings, factory=factory).execute(PROMPT)

    assert result.success is True
    assert result.processed_leads == 2
    assert browser.close_count == 1
    sheet = load_workbook(result.excel_output_path)["Leads"]
    assert sheet["B2"].value == "hello@good.example"
    assert sheet["B3"].value in (None, "")
