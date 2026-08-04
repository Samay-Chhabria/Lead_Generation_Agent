"""Requirement verification matrix (Requirements 1-14).

Each requirement from the assignment is verified here by exercising the real
component that implements it. The full end-to-end workflow runs through the
fake 'fixed' provider so no browser or network is used, except Requirement 3,
which launches the real browser automation.

Requirement matrix:
    R1  Accepts a natural-language prompt
    R2  Business category extracted from the prompt
    R3  Browser automation launches
    R4  Businesses are searched
    R5  Name, email, phone, website, and location are collected
    R6  Multiple businesses are collected
    R7  Missing fields are handled
    R8  An Excel workbook is generated
    R9  Required columns exist in the workbook
    R10 Output filename is meaningful
    R11 Execution summary is printed
    R12 Application runs without runtime errors
    R13 README instructions can be followed
    R14 Repository structure is complete
"""

from openpyxl import load_workbook

from app.browser.browser_manager import BrowserManager
from app.config.settings import Settings
from app.exporter.workbook_builder import COLUMN_HEADERS
from app.models.lead import Lead
from app.parser.prompt_parser import PromptParser
from app.pipeline.application_pipeline import ApplicationPipeline
from tests.conftest import PROJECT_ROOT

PROMPT = "software companies in Karachi"


def _lead(
    name: str = "Acme Corp",
    *,
    phone: str = "",
    email: str = "",
    website: str = "",
    location: str = "",
) -> Lead:
    return Lead(
        business_name=name,
        phone_number=phone,
        email=email,
        website=website,
        location=location,
        provider="fixed",
        search_query=PROMPT,
        source_url=f"https://maps.example/place/{name}",
    )


def _run(fixed_settings: Settings, fixed_factory, leads: list[Lead]):
    return ApplicationPipeline(settings=fixed_settings, factory=fixed_factory(leads)).execute(
        PROMPT
    )


def test_r1_accepts_natural_language_prompt(fixed_settings, fixed_factory) -> None:
    plan = PromptParser().parse(PROMPT, settings=fixed_settings)
    result = _run(fixed_settings, fixed_factory, [_lead()])

    assert plan.original_prompt == PROMPT
    assert result.success is True


def test_r2_business_category_extracted(fixed_settings) -> None:
    plan = PromptParser().parse("coffee shops in America", settings=fixed_settings)

    assert plan.business_type == "coffee shops"
    assert plan.location == "America"


def test_r3_browser_automation_launches(settings) -> None:
    manager = BrowserManager(settings=settings)
    try:
        page = manager.launch()
        assert manager.is_running()
        assert page is not None
    finally:
        manager.close()
    assert not manager.is_running()


def test_r4_businesses_are_searched(fixed_settings, fixed_factory) -> None:
    result = _run(fixed_settings, fixed_factory, [_lead("Alpha Corp"), _lead("Beta Ltd")])

    assert result.success is True
    assert result.collected_leads == 2


def test_r5_contact_fields_collected(fixed_settings, fixed_factory) -> None:
    result = _run(
        fixed_settings,
        fixed_factory,
        [
            _lead(
                "Acme Corp",
                phone="+12125551234",
                email="info@acme.example",
                website="https://acme.example",
                location="Karachi",
            )
        ],
    )

    sheet = load_workbook(result.excel_output_path)["Leads"]
    assert sheet["A2"].value == "Acme Corp"
    assert sheet["B2"].value == "info@acme.example"
    assert sheet["C2"].value == "+12125551234"
    assert sheet["D2"].value == "https://acme.example"
    assert sheet["E2"].value == "Karachi"


def test_r6_multiple_businesses_collected(fixed_settings, fixed_factory) -> None:
    leads = [_lead(name=f"Business {index}") for index in range(5)]

    result = _run(fixed_settings, fixed_factory, leads)

    assert result.success is True
    assert result.collected_leads == 5
    assert result.processed_leads == 5
    assert load_workbook(result.excel_output_path)["Leads"].max_row == 6


def test_r7_missing_fields_handled(fixed_settings, fixed_factory) -> None:
    result = _run(fixed_settings, fixed_factory, [_lead("Minimal Co"), _lead("No Data At All")])

    assert result.success is True
    sheet = load_workbook(result.excel_output_path)["Leads"]
    assert sheet["A2"].value == "Minimal Co"
    for column in ("B", "C", "D", "E"):
        assert sheet[f"{column}2"].value in (None, "")


def test_r8_excel_workbook_generated(fixed_settings, fixed_factory) -> None:
    result = _run(fixed_settings, fixed_factory, [_lead("Acme Corp")])

    assert result.excel_output_path is not None
    assert result.excel_output_path.exists()
    assert result.excel_output_path.suffix == ".xlsx"
    assert load_workbook(result.excel_output_path).sheetnames == ["Leads"]


def test_r9_required_columns_exist(fixed_settings, fixed_factory) -> None:
    result = _run(fixed_settings, fixed_factory, [_lead("Acme Corp")])

    sheet = load_workbook(result.excel_output_path)["Leads"]
    assert [cell.value for cell in sheet[1]] == COLUMN_HEADERS


def test_r10_output_filename_is_meaningful(fixed_settings, fixed_factory) -> None:
    result = _run(fixed_settings, fixed_factory, [_lead("Acme Corp")])

    assert result.excel_output_path is not None
    assert result.excel_output_path.name == "leads_software_companies_Karachi.xlsx"


def test_r11_execution_summary_printed(fixed_settings, fixed_factory, capsys) -> None:
    _run(fixed_settings, fixed_factory, [_lead("Acme Corp")])

    output = capsys.readouterr().out
    assert "Lead Generation Completed Successfully" in output
    assert "Search Query: software companies in Karachi" in output
    assert "Businesses Found: 1" in output
    assert "Leads Exported: 1" in output
    assert "Execution Time:" in output


def test_r12_runs_without_runtime_errors(fixed_settings, fixed_factory) -> None:
    successful = _run(fixed_settings, fixed_factory, [_lead("Acme Corp")])
    assert successful.success is True

    failed = _run(fixed_settings, fixed_factory, [])
    assert failed.success is True
    assert failed.excel_output_path is not None


def test_r13_readme_instructions_verified() -> None:
    readme = (PROJECT_ROOT / "README.md").read_text(encoding="utf-8")
    assert "## Installation" in readme
    assert "## Setup" in readme
    assert "## Running" in readme
    assert "python app/main.py" in readme
    assert "pytest" in readme

    env_template = (PROJECT_ROOT / ".env.example").read_text(encoding="utf-8")
    for variable in (
        "HEADLESS",
        "TIMEOUT",
        "MAX_LEADS",
        "SEARCH_PROVIDER",
        "OUTPUT_DIR",
        "LOG_DIR",
        "LOG_LEVEL",
    ):
        assert variable in env_template


def test_r14_repository_structure_verified() -> None:
    for path in (
        "app",
        "tests",
        "docs",
        "pyproject.toml",
        "requirements.txt",
        "README.md",
        ".env.example",
    ):
        assert (PROJECT_ROOT / path).exists(), f"Missing expected path: {path}"
