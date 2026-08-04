"""Integration tests for the end-to-end application pipeline.

The full workflow — prompt parsing, provider search, lead processing, Excel
export, and execution summary — is exercised against fake browsers and a
provider that hands the pipeline a fixed set of leads, so no network access or
real browser is ever used (Requirement 1, 11, 12).
"""

import logging
from dataclasses import replace
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.config.settings import Settings
from app.models.execution_result import ExecutionResult
from app.models.lead import Lead
from app.models.search_plan import SearchPlan
from app.pipeline.application_pipeline import ApplicationPipeline
from app.pipeline.search_pipeline import SearchPipeline
from app.providers.provider_factory import ProviderFactory
from app.providers.provider_registry import ProviderRegistry
from app.providers.search_provider import SearchProvider
from tests.fakes import FakeBrowser, FakeElement, FakePage

PROMPT = "software companies in Karachi"
MAILTO_SELECTOR = 'a[href^="mailto:"]'


@pytest.fixture
def settings(tmp_path: Path) -> Settings:
    return Settings(
        headless=True,
        timeout=2_000,
        max_leads=50,
        search_provider="fixed",
        browser_type="chromium",
        output_dir=tmp_path / "outputs",
        log_dir=tmp_path / "logs",
        log_level="INFO",
    )


class FixedLeadsProvider(SearchProvider):
    """A provider that hands the pipeline a fixed set of leads."""

    name = "fixed"
    current_leads: list[Lead] = []

    def __init__(self, browser, plan, settings, logger=None):
        super().__init__(browser=browser, plan=plan, settings=settings, logger=logger)
        self._page = browser.new_page()
        self._leads = list(FixedLeadsProvider.current_leads)

    @property
    def page(self):
        return self._page

    @property
    def leads(self):
        return self._leads

    def close(self) -> None:
        pass


def _lead(name: str = "Acme Corp", website: str = "") -> Lead:
    return Lead(
        business_name=name,
        phone_number="",
        email="",
        website=website,
        location="",
        provider="fixed",
        search_query="software companies in Karachi",
        source_url="https://maps.example/place/acme",
    )


def _factory(settings: Settings, browser: FakeBrowser) -> ProviderFactory:
    registry = ProviderRegistry()
    registry.register(FixedLeadsProvider)
    return ProviderFactory(registry=registry, settings=settings, browser=browser)


def _plan() -> SearchPlan:
    return SearchPlan(
        original_prompt=PROMPT,
        business_type="software companies",
        location="Karachi",
        provider="fixed",
        max_results=50,
    )


def _run(settings: Settings, browser: FakeBrowser) -> ExecutionResult:
    return ApplicationPipeline(settings=settings, factory=_factory(settings, browser)).execute(
        PROMPT
    )


# --- Complete execution ------------------------------------------------------


def test_complete_execution_collects_and_exports(settings: Settings, tmp_path: Path) -> None:
    browser = FakeBrowser()
    FixedLeadsProvider.current_leads = [_lead("Alpha Corp"), _lead("Beta Ltd"), _lead("Gamma Co")]

    result = _run(settings, browser)

    assert result.success is True
    assert result.search_query == "software companies in Karachi"
    assert result.business_type == "software companies"
    assert result.location == "Karachi"
    assert result.provider == "fixed"
    assert result.requested_leads == 50
    assert result.collected_leads == 3
    assert result.processed_leads == 3
    assert result.duplicates_removed == 0
    assert result.execution_time >= 0.0
    assert result.excel_output_path is not None
    assert result.excel_output_path.exists()

    workbook = load_workbook(result.excel_output_path)
    sheet = workbook["Leads"]
    assert sheet.max_row == 4
    assert sheet["A2"].value == "Alpha Corp"
    assert sheet["A3"].value == "Beta Ltd"
    assert sheet["A4"].value == "Gamma Co"


@pytest.mark.parametrize("count", [1, 50])
def test_execution_handles_small_and_large_searches(settings: Settings, count: int) -> None:
    browser = FakeBrowser()
    FixedLeadsProvider.current_leads = [_lead(name=f"Site {i}") for i in range(count)]

    result = _run(settings, browser)

    assert result.success is True
    assert result.collected_leads == count
    assert result.processed_leads == count


def test_execution_with_no_results_still_exports(settings: Settings) -> None:
    browser = FakeBrowser()
    FixedLeadsProvider.current_leads = []

    result = _run(settings, browser)

    assert result.success is True
    assert result.collected_leads == 0
    assert result.processed_leads == 0
    assert result.excel_output_path is not None
    assert result.excel_output_path.exists()
    assert load_workbook(result.excel_output_path)["Leads"].max_row == 1


def test_execution_continues_after_website_failures(settings: Settings) -> None:
    page = FakePage(
        elements_by_url={
            "https://good1.example": {
                MAILTO_SELECTOR: [FakeElement(attributes={"href": "mailto:one@good1.example"})]
            }
        },
        goto_errors={"https://broken.example": RuntimeError("connection refused")},
    )
    browser = FakeBrowser(page)
    FixedLeadsProvider.current_leads = [
        _lead(name="Good 1", website="https://good1.example"),
        _lead(name="Broken", website="https://broken.example"),
        _lead(name="Good 2", website="https://good2.example"),
    ]

    result = _run(settings, browser)

    assert result.success is True
    assert result.collected_leads == 3
    assert result.processed_leads == 3
    assert result.excel_output_path is not None


def test_execution_export_failure_returns_unsuccessful(settings: Settings, tmp_path: Path) -> None:
    blocker = tmp_path / "outputs"
    blocker.write_bytes(b"not a directory")
    blocked = replace(settings, output_dir=blocker)
    browser = FakeBrowser()
    FixedLeadsProvider.current_leads = [_lead("Alpha Corp")]

    result = ApplicationPipeline(settings=blocked, factory=_factory(blocked, browser)).execute(
        PROMPT
    )

    assert result.success is False
    assert result.business_type == "software companies"
    assert result.excel_output_path is None


def test_execution_invalid_prompt_returns_unsuccessful(settings: Settings) -> None:
    browser = FakeBrowser()

    result = ApplicationPipeline(settings=settings, factory=_factory(settings, browser)).execute(
        "just a phrase"
    )

    assert result.success is False
    assert result.excel_output_path is None
    assert browser.launch_count == 0


# --- Console output ----------------------------------------------------------


def test_execution_prints_plan_and_summary(
    settings: Settings, capsys: pytest.CaptureFixture
) -> None:
    browser = FakeBrowser()
    FixedLeadsProvider.current_leads = [_lead("Alpha Corp")]

    _run(settings, browser)

    output = capsys.readouterr().out
    assert "Search Plan" in output
    assert "Original Prompt: software companies in Karachi" in output
    assert "Lead Generation Completed Successfully" in output
    assert "Search Query: software companies in Karachi" in output
    assert "Business Type: software companies" in output
    assert "Location: Karachi" in output
    assert "Businesses Found: 1" in output
    assert "Leads Exported: 1" in output
    assert "Execution Time:" in output


def test_execution_failure_prints_failure_summary(
    settings: Settings, capsys: pytest.CaptureFixture
) -> None:
    browser = FakeBrowser()
    FixedLeadsProvider.current_leads = []

    ApplicationPipeline(settings=settings, factory=_factory(settings, browser)).execute(
        "no location here"
    )

    output = capsys.readouterr().out
    assert "Lead Generation Failed" in output
    assert "Search Query: no location here" in output


# --- Lifecycle and logging ---------------------------------------------------


def test_execution_closes_browser(settings: Settings) -> None:
    browser = FakeBrowser()
    FixedLeadsProvider.current_leads = [_lead("Alpha Corp")]

    _run(settings, browser)

    assert browser.close_count == 1


def test_execution_logs_every_stage(settings: Settings, caplog: pytest.LogCaptureFixture) -> None:
    browser = FakeBrowser()
    FixedLeadsProvider.current_leads = [_lead("Alpha Corp")]

    with caplog.at_level(logging.INFO):
        _run(settings, browser)

    messages = [record.message for record in caplog.records]
    for stage in (
        "Pipeline started.",
        "Parsed search plan:",
        "Browser started.",
        "Provider initialized.",
        "Processing completed.",
        "Excel exported.",
        "Summary generated.",
        "Pipeline finished.",
    ):
        assert any(stage in message for message in messages)


# --- SearchPipeline.run_and_export -------------------------------------------


def test_run_and_export_returns_three_artifacts(settings: Settings) -> None:
    browser = FakeBrowser()
    FixedLeadsProvider.current_leads = [_lead("Alpha Corp")]
    pipeline = SearchPipeline(factory=_factory(settings, browser))

    provider_result, processing_result, path = pipeline.run_and_export(_plan())

    assert provider_result.success is True
    assert provider_result.lead_count == 1
    assert processing_result.final_count == 1
    assert path.exists()
