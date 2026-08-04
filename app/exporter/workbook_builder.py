"""Excel workbook construction.

WorkbookBuilder turns processed Lead objects into a formatted .xlsx workbook.
It owns the spreadsheet layout: a single ``Leads`` sheet, a bold header row, a
frozen header pane, sensible column widths, and UTF-8-safe cell values. The
builder only renders already-processed data — it never validates or changes the
leads it receives (Requirement 8, Requirement 9).
"""

import logging
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.config.logging_config import get_logger
from app.models.lead import Lead

SHEET_NAME = "Leads"
COLUMN_HEADERS = [
    "Business Name",
    "Email",
    "Phone Number",
    "Website",
    "Location",
    "Provider",
    "Search Query",
    "Collected At",
    "Source URL",
]
_COLUMN_PADDING = 2
_MAX_COLUMN_WIDTH = 60


class WorkbookBuilder:
    """Render a list of leads into an openpyxl Workbook."""

    def __init__(self, logger: logging.Logger | None = None) -> None:
        """Initialize the builder with an optional logger."""
        self._logger = logger or get_logger("exporter")

    def build(self, leads: list[Lead]) -> Workbook:
        """Build a formatted workbook containing the given leads.

        Args:
            leads: The processed leads to write. Empty strings are written as
                blank cells.

        Returns:
            A ready-to-save openpyxl Workbook.
        """
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = SHEET_NAME

        sheet.append(COLUMN_HEADERS)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"

        rows = [self._to_row(lead) for lead in leads]
        for row in rows:
            sheet.append(row)

        self._autosize(sheet, rows)

        self._logger.info("Workbook created.")
        self._logger.info("Rows written: %d.", len(rows))
        return workbook

    @staticmethod
    def _to_row(lead: Lead) -> list[str]:
        """Convert a lead into the header-aligned row of cell values."""
        collected = lead.collected_at
        if isinstance(collected, datetime):
            collected_at = collected.strftime("%Y-%m-%d %H:%M:%S")
        else:
            collected_at = str(collected)
        return [
            lead.business_name,
            lead.email,
            lead.phone_number,
            lead.website,
            lead.location,
            lead.provider,
            lead.search_query,
            collected_at,
            lead.source_url,
        ]

    @staticmethod
    def _autosize(sheet: Worksheet, rows: list[list[str]]) -> None:
        """Size each column to its widest value, honoring the headers."""
        for index, header in enumerate(COLUMN_HEADERS):
            widest = len(header)
            for row in rows:
                widest = max(widest, len(row[index]))
            width = min(widest + _COLUMN_PADDING, _MAX_COLUMN_WIDTH)
            sheet.column_dimensions[get_column_letter(index + 1)].width = width
